from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def get_styles():
    bold_font = Font(bold=True)
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    alignment_center = Alignment(horizontal="center")
    alignment_right = Alignment(horizontal="right")
    alignment_left = Alignment(horizontal="left")
    return {
        "bold_font": bold_font,
        "black_fill": black_fill,
        "white_fill": white_fill,
        "red_fill": red_fill,
        "alignment_center": alignment_center,
        "alignment_right": alignment_right,
        "alignment_left": alignment_left,
        "thin_border": thin_border,
    }
