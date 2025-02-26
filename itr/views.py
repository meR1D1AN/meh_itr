# import urllib.parse
from datetime import datetime

from django import forms
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db.models import Q, QuerySet
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import CreateView, DetailView, ListView, UpdateView
from openpyxl import Workbook

# from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from itr.excel.data_utils import get_month_days, get_workdays
from itr.excel.filename import get_filename
from itr.excel.get_current_month_year import get_current_month_year
from itr.excel.styles import get_styles
from itr.forms import CombinedEmployeeCustomerForm, CustomerForm, VacationForm
from itr.models import Customer, Employee, Vacation, VacationStatus

from .forms import MONTHS


# from itr.salary import calculate_salary
# from itr.utils import sanitize_filename


class EmployeeListView(LoginRequiredMixin, ListView):
    model = Employee
    template_name = "itr/employee_list.html"
    context_object_name = "employees"

    def get_queryset(self):
        # Если пользователь — администратор, показываем всех сотрудников
        if self.request.user.is_superuser:
            return Employee.objects.all().order_by("last_name")
        # Если обычный пользователь, показываем только тех, кого он сам создал
        else:
            queryset = Employee.objects.filter(created_by=self.request.user).order_by("last_name")
        return queryset.prefetch_related("customer")


class EmployeeCreateView(LoginRequiredMixin, CreateView):
    model = Employee
    form_class = CombinedEmployeeCustomerForm
    template_name = "itr/employee_form.html"
    success_url = reverse_lazy("itr:employee_list")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user  # Передаем текущего пользователя в форму
        return kwargs

    def form_valid(self, form):
        customers = []
        # Создаем или получаем существующего заказчика
        customer = form.cleaned_data.get("customer")
        if customer:
            if isinstance(customer, QuerySet):
                customers = list(customer)
            else:
                customers = [customer]
        else:
            customer_form = CustomerForm(self.request.POST)
            if customer_form.is_valid():
                customer = customer_form.save(commit=False)
                customer.created_by = self.request.user
                customer.save()
                customers = [customer]
            else:
                return self.form_invalid(form)

        # Сначала сохраняем объект Employee
        employee = form.save(commit=False)  # Не сохраняем в БД, чтобы получить id
        employee.created_by = self.request.user
        employee.save()  # Теперь сохраняем объект в БД
        form.save_m2m()  # Сохраняем ManyToMany связи

        # Привязка заказчиков
        employee.customer.set(customers)

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["is_update"] = False  # Явно указываем, что это создание
        return context

    def form_invalid(self, form):
        print(form.errors)  # Вывод ошибок в консоль
        return super().form_invalid(form)


class EmployeeDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Employee
    template_name = "itr/employee_detail.html"
    context_object_name = "employee"

    def get_object(self, queryset=None):
        return Employee.objects.prefetch_related("customer").get(pk=self.kwargs["pk"])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        employee = self.get_object()
        context["customer"] = employee.customer
        return context

    def test_func(self):
        employee = self.get_object()
        return self.request.user.is_superuser or self.request.user == employee.created_by


class EmployeeUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Employee
    form_class = CombinedEmployeeCustomerForm
    template_name = "itr/employee_form.html"
    success_url = reverse_lazy("itr:employee_list")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user  # Передаем текущего пользователя в форму
        return kwargs

    def form_valid(self, form):
        form.instance.created_by = self.request.user  # Автоматически назначаем создателя
        return super().form_valid(form)

    def test_func(self):
        employee = self.get_object()
        return self.request.user.is_superuser or self.request.user == employee.created_by

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["is_update"] = True
        context["employee"] = self.object
        return context


class CustomerCreateView(LoginRequiredMixin, CreateView):
    model = Customer
    form_class = CustomerForm
    template_name = "itr/customer_form.html"
    success_url = reverse_lazy("itr:customer_list")

    def form_valid(self, form):
        form.instance.created_by = self.request.user  # Назначаем создателя
        self.object = form.save()

        print(f"Создан заказчик: {self.object.customer_name}, ID: {self.object.id}")

        # Проверяем параметр 'redirect' в запросе
        redirect_param = self.request.GET.get("redirect")
        if redirect_param == "list":
            # Перенаправляем на список заказчиков
            return HttpResponseRedirect(self.success_url)
        else:
            # Возвращаем JS-код для закрытия окна и обновления списка заказчиков
            return HttpResponse(
                f"<script>window.opener.addNewCustomerOption({self.object.id}, '{self.object.customer_name}'); "
                f"window.close();</script>"
            )


class CustomerListView(LoginRequiredMixin, ListView):
    model = Customer
    template_name = "itr/customer_list.html"
    context_object_name = "customers"

    def get_queryset(self):
        # Если пользователь — администратор, показываем всех сотрудников
        if self.request.user.is_superuser:
            return Customer.objects.all().order_by("customer_name")
        # Если обычный пользователь, показываем только тех, кого он сам создал
        return Customer.objects.filter(created_by=self.request.user).order_by("customer_name")


class CustomerUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Customer
    form_class = CustomerForm
    template_name = "itr/customer_form.html"
    success_url = reverse_lazy("itr:customer_list")

    def form_valid(self, form):
        form.instance.created_by = self.request.user  # Автоматически назначаем создателя
        return super().form_valid(form)

    def test_func(self):
        customer = self.get_object()
        return self.request.user.is_superuser or self.request.user == customer.created_by

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["is_update"] = True
        context["customer"] = self.get_object()
        return context


class CustomerDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Customer
    template_name = "itr/customer_detail.html"
    context_object_name = "customer"

    def test_func(self):
        customer = self.get_object()
        return self.request.user.is_superuser or self.request.user == customer.created_by


class CustomerExcelView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        customer = get_object_or_404(Customer, pk=self.kwargs["pk"])
        return self.request.user.is_superuser or self.request.user == customer.created_by

    def get(self, request, pk):
        # получаем список заказчиков
        customer = get_object_or_404(Customer, pk=pk)

        # Определяем месяц и год
        current_month, current_year = get_current_month_year(request)

        # Определяем название месяца
        month_name_ru = MONTHS[current_month]

        # Получаем количество дней в текущем месяце
        days_in_month = get_month_days(current_month, current_year)

        # сохранение файла в формате
        encoded_filename = get_filename(customer, month_name_ru, current_year)

        # создаём книгу и лист
        wb = Workbook()
        ws = wb.active
        ws.title = f"График работы {customer.customer_name}"

        styles = get_styles()

        # Заголовок
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
        ws.cell(
            row=1,
            column=1,
            value="Табель учета использования рабочего времени лифтеров-проводников",
        ).font = styles["bold_font"]
        # График работы
        ws.cell(row=2, column=1, value=f"График {customer.work_schedule}").font = styles["bold_font"]
        # Адрес
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=20)
        ws.cell(row=3, column=1, value=f"Адрес: {customer.address_customer}").font = styles["bold_font"]
        # Месяц и год
        headers_row_y = 5
        ws.cell(
            row=headers_row_y,
            column=1,
            value=f"{month_name_ru.capitalize()} {current_year}",
        ).font = styles["bold_font"]
        ws.cell(row=headers_row_y, column=1).border = styles["thin_border"]
        # Заголовки столбцов
        headers_row = 6
        ws.cell(row=headers_row, column=1, value="Ф.И.О.").font = styles["bold_font"]
        ws.cell(row=headers_row, column=1).border = styles["thin_border"]

        # Добавляем дни месяца в заголовок
        for day in range(1, days_in_month + 1):
            cell = ws.cell(row=headers_row, column=day + 1, value=day)
            cell.font = styles["bold_font"]
            cell.border = styles["thin_border"]
            cell.alignment = styles["alignment_center"]

            if datetime(current_year, current_month, day).weekday() >= 5:
                cell.fill = styles["red_fill"]

        # Добавляем "Всего смен"
        total_shifts_col = days_in_month + 2
        ws.cell(row=headers_row, column=total_shifts_col, value="Всего смен").font = styles["bold_font"]
        ws.cell(row=headers_row, column=total_shifts_col).border = styles["thin_border"]
        ws.cell(row=headers_row, column=total_shifts_col).alignment = styles["alignment_center"]

        # Группируем рабочие дни по сотрудникам
        employee_workdays = get_workdays(customer, current_month, current_year)
        total_shifts_count = 0

        # Заполняем данные
        for row, (employee, work_days) in enumerate(employee_workdays.items(), start=7):
            # ФИО сотрудника
            ws.cell(row=row, column=1, value=str(employee)).border = styles["thin_border"]

            # Заполнение дней
            for day in range(1, days_in_month + 1):
                cell = ws.cell(row=row, column=day + 1, value="")
                cell.border = styles["thin_border"]

                # Если день есть в рабочих днях - закрашиваем черным, иначе белым
                if day in work_days:
                    cell.fill = styles["black_fill"]
                    total_shifts_count += 1
                else:
                    cell.fill = styles["white_fill"]

            # Записываем общее количество смен
            ws.cell(row=row, column=total_shifts_col, value=len(work_days)).border = styles["thin_border"]
            ws.cell(row=row, column=total_shifts_col, value=len(work_days)).alignment = styles["alignment_center"]

        # Добавление итоговой строки ПОСЛЕ завершения цикла
        total_row = row + 1

        # Вместо объединения - просто создаем ячейку с надписью
        ws.merge_cells(
            start_row=total_row,
            start_column=1,
            end_row=total_row,
            end_column=days_in_month + 1,
        )
        ws.cell(row=total_row, column=1, value="Всего смен:").font = styles["bold_font"]
        ws.cell(row=total_row, column=1).alignment = styles["alignment_right"]

        # Записываем общее количество смен в ячейку, которая не является частью объединенной области
        total_shifts = sum(len(work_days) for _, work_days in employee_workdays.items())
        ws.cell(row=total_row, column=total_shifts_col, value=total_shifts).font = styles["bold_font"]
        ws.cell(row=total_row, column=total_shifts_col).border = styles["thin_border"]
        ws.cell(row=total_row, column=total_shifts_col).alignment = styles["alignment_center"]

        # Автоширина столбцов
        for col in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            if col == 1:
                for employee, _ in employee_workdays.items():
                    employee_name = f"{employee.last_name} {employee.first_name}"
                    if len(employee_name) > max_length:
                        max_length = len(employee_name)
            else:
                for row in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col)
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (TypeError, ValueError):
                        pass
            ws.column_dimensions[column_letter].width = max_length + 2

        # Делаем подпись согласно фирме в классе Customer
        total_row = row + 4
        ws.merge_cells(range_string="F{}:P{}".format(total_row, total_row))
        ws.cell(row=total_row, column=6, value="Подрядчик:")
        ws.merge_cells(range_string="S{}:AB{}".format(total_row, total_row))
        ws.cell(row=total_row, column=19, value="Заказчик:")

        total_row = row + 5
        ws.merge_cells(range_string="F{}:P{}".format(total_row, total_row))
        if customer.firm == 'ООО ГК "ЛифтСервис"':
            ws.cell(row=total_row, column=6, value="Управляющий")
        else:
            ws.cell(row=total_row, column=6, value="Генеральный директор")

        total_row = row + 6
        ws.merge_cells(range_string="F{}:P{}".format(total_row, total_row))
        if customer.firm == 'ООО ГК "ЛифтСервис"':
            ws.cell(row=total_row, column=6, value='ООО ГК "ЛифтСервис"')
        elif customer.firm == 'ООО "ЦентрСервис"':
            ws.cell(row=total_row, column=6, value='ООО "ЦентрСервис"')
        elif customer.firm == 'ООО Холдинг "ЦентрСервис"':
            ws.cell(row=total_row, column=6, value='ООО Холдинг "ЦентрСервис"')
        ws.merge_cells(range_string="S{}:AB{}".format(total_row, total_row))
        ws.cell(row=total_row, column=19, value=f"{customer.customer_name}")

        total_row = row + 8
        ws.merge_cells(range_string="F{}:P{}".format(total_row, total_row))
        if customer.firm == 'ООО ГК "ЛифтСервис"':
            ws.cell(row=total_row, column=6, value="___________________ Г. А. Давбер")
        elif customer.firm == 'ООО "ЦентрСервис"':
            ws.cell(row=total_row, column=6, value="___________________ Д. И. Сычёв")
        elif customer.firm == 'ООО Холдинг "ЦентрСервис"':
            ws.cell(row=total_row, column=6, value="___________________ Д. В. Романов")
        ws.merge_cells(range_string="S{}:AB{}".format(total_row, total_row))
        ws.cell(row=total_row, column=19, value="___________________ /________________/")

        # Сохраняем файл для скачивания
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f"attachment; filename*=UTF-8''{encoded_filename}"

        wb.save(response)
        return response


def search_view(request):
    query = request.GET.get("query", "")
    employee_results = Employee.objects.filter(
        Q(last_name__icontains=query)
        | Q(first_name__icontains=query)
        | Q(middle_name__icontains=query)
        | Q(metro__icontains=query)
        | Q(firm__icontains=query)
        | Q(residential_address__icontains=query)
        | Q(phone__icontains=query)
    )
    customer_results = Customer.objects.filter(
        Q(customer_name__icontains=query)
        | Q(address_customer__icontains=query)
        | Q(work_schedule__icontains=query)
        | Q(responsible_person__icontains=query)
        | Q(salary__icontains=query)
        | Q(firm__icontains=query)
    )
    context = {
        "employee_results": employee_results,
        "customer_results": customer_results,
        "query": query,
    }
    return render(request, "itr/search_results.html", context)


class VacationListView(LoginRequiredMixin, ListView):
    """
    Список отпусков с разными правами доступа
    """

    model = Vacation
    template_name = "itr/vacation_list.html"
    context_object_name = "vacations"

    def get_queryset(self):
        # Для суперпользователя - все отпуска
        if self.request.user.is_superuser:
            return Vacation.objects.all()

        # Для обычного пользователя - только его собственные
        return Vacation.objects.filter(user=self.request.user)


class VacationCreateView(LoginRequiredMixin, CreateView):
    """
    Создание отпуска
    """

    model = Vacation
    form_class = VacationForm
    template_name = "itr/vacation_form.html"
    success_url = reverse_lazy("itr:vacation_list")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["initial"] = {"user": self.request.user}
        return kwargs

    def form_valid(self, form):
        # Автоматически устанавливаем текущего пользователя
        form.instance.user = self.request.user
        # По умолчанию статус - на согласовании
        form.instance.status = VacationStatus.PENDING
        return super().form_valid(form)


class VacationUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """
    Редактирование отпуска с проверкой прав доступа
    """

    model = Vacation
    form_class = VacationForm
    template_name = "itr/vacation_form.html"
    success_url = reverse_lazy("itr:vacation_list")

    def handle_no_permission(self):
        """
        Переопределяем обработку отсутствия прав доступа
        """
        vacation = self.get_object()
        messages.error(
            self.request,
            "Согласованный отпуск может отредактировать только администратор.",
        )
        return redirect("itr:vacation_detail", pk=vacation.pk)

    def test_func(self):
        """
        Проверка прав доступа:
        - Только создатель может редактировать
        - Администратор может редактировать любые отпуска
        - Нельзя редактировать согласованные отпуска
        """
        vacation = self.get_object()
        return self.request.user.is_superuser or (
            vacation.user == self.request.user and vacation.status != VacationStatus.APPROVED
        )

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["initial"] = {"user": self.object.user or self.request.user}
        return kwargs

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Если сотрудник не выбран, скрываем поле employee
        if not self.object.employee:
            form.fields["employee"].widget = forms.HiddenInput()
        return form

    def form_valid(self, form):
        # Если employee не выбран, устанавливаем текущего пользователя
        if not form.cleaned_data.get("employee"):
            form.instance.user = self.request.user

        # Если не администратор, всегда устанавливаем статус "на согласовании"
        if not self.request.user.is_superuser:
            form.instance.status = VacationStatus.PENDING

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        """
        Добавляем информацию о текущем статусе в контекст
        """
        context = super().get_context_data(**kwargs)
        context["current_status"] = self.object.status
        context["is_update"] = True
        return context


class VacationDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    """
    Детальный просмотр отпуска с проверкой прав
    """

    model = Vacation
    template_name = "itr/vacation_detail.html"
    context_object_name = "vacation"

    def test_func(self):
        # Проверка прав доступа
        vacation = self.get_object()
        return self.request.user.is_superuser or vacation.user == self.request.user

    # @register.filter
    # def get_days_text(days):
    #     if days % 10 == 1 and days % 100 != 11:
    #         return f"{days} день"
    #     elif 2 <= days % 10 <= 4 and (days % 100 < 10 or days % 100 >= 20):
    #         return f"{days} дня"
    #     else:
    #         return f"{days} дней"


class VacationApproveView(LoginRequiredMixin, UserPassesTestMixin, View):
    """
    Представление для согласования отпуска (не generic)
    """

    def test_func(self):
        # Только суперпользователь может согласовывать
        return self.request.user.is_superuser

    def post(self, request, pk):
        vacation = get_object_or_404(Vacation, pk=pk)

        # Получаем статус из формы или параметров
        status = request.POST.get("status", VacationStatus.APPROVED)

        if status in dict(VacationStatus.choices):
            vacation.status = status
            vacation.itr = request.user  # Кто согласовал
            vacation.save()

        # Перенаправляем на страницу деталей отпуска
        return redirect("itr:vacation_detail", pk=pk)
